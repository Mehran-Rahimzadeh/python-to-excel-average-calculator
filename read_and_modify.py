# clearing the terminal befor each run
import os
clear=lambda: os.system("clear")
clear()

#-------------------------Loading Exell file for modifying-------------

import openpyxl 
from openpyxl import Workbook
from openpyxl import load_workbook


    
def average():
    import openpyxl 
    from openpyxl import Workbook
    from openpyxl import load_workbook

    
        # workbook object is created
    path = "example_op.xlsx"
    wb_obj = openpyxl.load_workbook(path)

    # Get workbook active sheet object
    sheet_obj = wb_obj.active
    sheet_obj.cell(row=1, column=6).value = "Average"
   
#  converting credits from string in to list of integers--
    for s in range (2,s):
        def str_list_credits():
            cell_obj = sheet_obj.cell(row = s, column = 4)
            credit_list=[]
            # Print value of cell object
            # using the value attribute
            co = cell_obj.value
            b=co.replace(" ","")
            bb= list((b.replace(",","")))
            credit_list=[]
            for i in range (len(bb)):
                credit_list.append(int(bb[i]))
            return credit_list

        #  converting scores from string in to list of integers--
        def str_listt_scores():
                cell_obj = sheet_obj.cell(row = s, column = 5)
                score_list=[]
                # using the value attribute
                co = cell_obj.value
                k=co.replace(" ","")
                kk= list((k.replace(",","" "")))
                kkk=[]
                score_list=[]
                for i in range (len(kk)):
                    kkk.append(int(kk[i]))
                
                for j in range (((len(kkk)-1)//2)+1):
                    t=2*j            
                    score_list.append(kkk[t]*10+kkk[t+1])
                return score_list
        # ----- calculating average
        def average_calculator():
            x=str_list_credits()
            y=str_listt_scores()
            m=[]
            n=sum(x)
            for i in range (len(x)):
               m.append(x[i]*y[i])
            average= (sum(m))/n
            average=round(average,2)
            return average
    
        #  -----puting results in Exell file--------
        def print_result():
            sheet_obj.cell(row= s, column=6 ).value = average_calculator()
            wb_obj.save(path)
    
        print_result()

average()


     
     


