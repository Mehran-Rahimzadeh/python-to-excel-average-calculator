# clearing the terminal befor each run
import os
clear=lambda: os.system("clear")
clear()
# ---------------------

import openpyxl 
from openpyxl import Workbook
from openpyxl import load_workbook

# creat empty list for student names and ther cours 
class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
data=[]
student=0
number = 3
co=[]
#asking funxtion
def ask():
    globals()["student"]= dict(firstName="" , lastName = "",
                                courses=[], credits=[], scores=[])
    name_ask()
    global co
    co =[]
    courses_ask()
    
    next_step_student()


# ---------getting first and last name--------

def name_ask():
    
    student["firstName"]=input("\nEnter the student name: ")
    student["lastName"]=input ("\nEnter the student last name: ")
    

#--------------getting courses---------
def courses_ask():
    global co
    cou=input ("\nEnter the student course: ")
    if cou in co:
        # print(" You have inseted this cours befor!")
        print(bcolors.WARNING + "Warning: You have inseted this cours before!" + bcolors.ENDC)
    else: 
        co.append(cou)
    
        student["courses"].append(cou)
        student["credits"].append(int(input ("\nEnter the course credit: ")))
        student["scores"].append(int(input ("\nEnter the course score: ")))
    next_step_courses()




 # input method (courses)
def next_step_courses():
    answer=ord(input('\nIs there any other courses? y/n: '))
    if answer==121:
        courses_ask()
    else:
        return
    


    # input method (courses)
def next_step_student():
   
    global number
    student["courses"]=( ", ".join( repr(e) for e in student["courses"]) )
    student["credits"]=( ", ".join( repr(e) for e in student["credits"] ) )
    student["scores"]=( ", ".join( repr(e) for e in student["scores"] ) )
    answer=ord(input('\nIs there any other students? y/n: '))
    if answer==121:
        data.append(student)
        number= number +1
        print(bcolors.OKBLUE + "\n----------------------------------" + bcolors.ENDC)
        ask()
    else:
        data.append(student)
    answer_print=ord(input('\nDo you want to have data as an Exelle file('+ bcolors.OKGREEN+' ("data.xlsx")'+ bcolors.ENDC+') ? y/n: '))
    if answer_print==121:
        pass
    else:
        exit()
        # print(*data, sep="\n")

ask()
# -----------------------make_exell()-------------------------------------------------------------
# -----------------------make_exell()-------------------------------------------------------------


row = 0
col = 0
wb = Workbook()
ws = wb.active

def create_xls(filepath):
    wb = Workbook()
    wb.save(filepath)
    from openpyxl.styles import Font
    ft = Font(bold=True)
    for row in ws["A1:E1"]:
        for cell in row:
            cell.font = ft

def write_xls(filepath, dictionary):
    wb = load_workbook(filepath)
    ws = wb.active
   
    headers = [x for x in dictionary[0]]
    for index, value in enumerate(headers):
         ws.cell(row=1, column=index+1).value = value


    for i, x in enumerate(dictionary):
        for idx,value in enumerate(x.values()):
            ws.cell(row=i+2, column=idx+1).value = value
    wb.save(filepath)
ws.cell(row=25, column=1).value = len(data)
create_xls("data.xlsx")
write_xls("data.xlsx", data)

# ----------sheet style-----
from openpyxl.styles import Font
ft = Font(bold=True)
for row in ws["A1:E1"]:
    for cell in row:
        cell.font = ft






# ---------calculating average from Data  and put the result in Exell file -----
# -------------------------------------------------------------------------------

    # workbook object is created
data_path = "./data.xlsx"
result_path= "./result.xlsx"
wb_obj = openpyxl.load_workbook(data_path)

# Get workbook active sheet object
sheet_obj = wb_obj.active
sheet_obj.cell(row=1, column=7).value = "Average"
   
#  converting credits from string in to list of integers--
    
def str_list_credits(s):
    
    cell_obj = sheet_obj.cell(row = s, column = 4)
    credit_list=[]
    # Print value of cell object
    # using the value attribute
    co = cell_obj.value
    b=co.replace(" ","")
    bb= list((b.replace(",","")))
    credit_list=[]
    for i in range (len(bb)):
        credit_list.append(int(bb[i]))
    return credit_list

#  converting scores from string in to list of integers--
def str_listt_scores(s):
    
    cell_obj = sheet_obj.cell(row = s, column = 5)
    score_list=[]
    # using the value attribute
    co = cell_obj.value
    k=co.replace(" ","")
    kk= list((k.replace(",","" "")))
    kkk=[]
    score_list=[]
    for i in range (len(kk)):
        kkk.append(int(kk[i]))
    
    for j in range (((len(kkk)-1)//2)+1):
        t=2*j            
        score_list.append(kkk[t]*10+kkk[t+1])
    return score_list
# ----- calculating average
def average_calculator(s):
    
    x=str_list_credits(s)
    y=str_listt_scores(s)
    m =[]
    n=sum(x)
    for i in range (len(x)):
        m.append(x[i]*y[i])
    average= float((sum(m))/n)
    average=round(average,2)

    sheet_obj.cell(row= s, column=7 ).value = average
    wb_obj.save(result_path)
 

def average(s):
    
    for s in range (2, s):
         str_list_credits(s)
         str_listt_scores(s)
         average_calculator(s)


calc=ord(input('\nDo you want to calculat the average\n  and report as an Exelle file' + bcolors.OKGREEN +' ("result.xlsx")'+ bcolors.ENDC+') ? y/n: '))
if calc==121:
    average(number)
else: 
    exit()