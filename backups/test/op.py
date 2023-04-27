import openpyxl 
from openpyxl import Workbook

data = [
    {'Name': 'John', 'Age': str([1,3,6]), 'Gender': 'Male'},
    {'Name': 'Sarah', 'Age': 4, 'Gender': 'Female'},
    {'Name': 'Michael', 'Age': 27, 'Gender': 'Male'},
]

from openpyxl import load_workbook

row = 0
col = 0

def create_xls(filepath):
    wb = Workbook()
    wb.save(filepath)

def write_xls(filepath, dictionary):
    wb = load_workbook(filepath)
    ws = wb.active

    headers = [x for x in dictionary[0]]
    for index, value in enumerate(headers):
         ws.cell(row=1, column=index+1).value = value


    for i, x in enumerate(dictionary):
        for idx,value in enumerate(x.values()):
            ws.cell(row=i+2, column=idx+1).value = value
    wb.save(filepath)

create_xls("example_op.xlsx")
write_xls("example_op.xlsx", data)




