# clearing the terminal befor each run
import os
clear=lambda: os.system("clear")
clear()
# ---------------------


# creat empty list for student names and ther cours 
data=[]
student=0


#asking funxtion
def ask():
    globals()["student"]= dict(firstName="" , lastName = "",
                                courses=[], credits=[], scores=[])
    name_ask()
    courses_ask()
    next_step_student()


# ---------getting first and last name--------
def name_ask():
    student["firstName"]=input("Enter the student name: ")
    student["lastName"]=input ("Enter the student last name: ")


#--------------getting courses---------
def courses_ask():
    student["courses"].append(input ("Enter the student course: "))
    student["credits"].append(int(input ("Enter the course credit: ")))
    student["scores"].append(float(input ("Enter the course scores: ")))
    next_step_courses()




 # input method (courses)
def next_step_courses():
    answer=ord(input('Is there any other courses? y/n: '))
    if answer==121:
        courses_ask()
    else:
        return
    


    # input method (courses)
def next_step_student():
    # student["courses"]=str(student["courses"])
    # student["credits"]=str(student["credits"])
    # student["scores"]=str(student["scores"])

    student["courses"]=( ", ".join( repr(e) for e in student["courses"]) )
    student["credits"]=( ", ".join( repr(e) for e in student["credits"] ) )
    student["scores"]=( ", ".join( repr(e) for e in student["scores"] ) )
    answer=ord(input('Is there any other students? y/n: '))
    if answer==121:
        data.append(student)
        ask()
    else:
        data.append(student)
        print(*data, sep="\n")

ask()
# make_title_exell()

import openpyxl 
from openpyxl import Workbook
from openpyxl import load_workbook

row = 0
col = 0
wb = Workbook()
ws = wb.active

def create_xls(filepath):
    wb = Workbook()
    wb.save(filepath)
    from openpyxl.styles import Font
    ft = Font(bold=True)
    for row in ws["A1:E1"]:
        for cell in row:
            cell.font = ft

def write_xls(filepath, dictionary):
    wb = load_workbook(filepath)
    ws = wb.active

    headers = [x for x in dictionary[0]]
    for index, value in enumerate(headers):
         ws.cell(row=1, column=index+1).value = value


    for i, x in enumerate(dictionary):
        for idx,value in enumerate(x.values()):
            ws.cell(row=i+2, column=idx+1).value = value
    wb.save(filepath)

create_xls("example_op.xlsx")
write_xls("example_op.xlsx", data)

# ----------sheet style-----
from openpyxl.styles import Font
ft = Font(bold=True)
for row in ws["A1:E1"]:
    for cell in row:
        cell.font = ft





#----------------------------Writing Data in to an Exell file-----
#                --------------------------------------------------
#                              -------------------------------


# #Create an new Excel file and add a worksheet.
# workbook=xlsxwriter.Workbook("karnameh_1.xlsx")
# worksheet = workbook.add_worksheet()

# # Widen the first column to make the text clearer.
# worksheet.set_column("A:A", 10)
# worksheet.set_column("B:B", 10)
# worksheet.set_column("C:C", 40)
# worksheet.set_column("D:D", 20)
# worksheet.set_column("E:E", 20)

# # Add a bold format to use to highlight cells.
# bold = workbook.add_format({"bold":True})


# # Write some header.
# worksheet.write('A1', "First Name", bold)
# worksheet.write('B1', 'Last Name', bold)
# worksheet.write('C1', 'Courses', bold)
# worksheet.write('D1', 'Credits', bold)
# worksheet.write('E1', 'Scores', bold)


# workbook.close()



# ---------------------------------Inserting Data in Exell file and make version 2-------------------


# from openpyxl import Workbook
# #wb = Workbook()

# # open workbook
# from openpyxl import load_workbook
# wb = load_workbook(filename = 'karnameh_1.xlsx')